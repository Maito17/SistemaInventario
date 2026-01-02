#cliente/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from .models import Cliente
from .forms import ClienteForm, ExcelUploadForm
import openpyxl
from django.http import HttpResponse
from decimal import Decimal

@login_required
def exportar_clientes_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ['ID Cliente', 'Nombre', 'Apellido', 'RUC/Cédula', 'Email', 'Teléfono', 'Dirección', 'Crédito Activo', 'Límite Crédito', 'Días Plazo']
    ws.append(headers)

    clientes = Cliente.objects.filter(user=request.user)
    for c in clientes:
        ws.append([
            c.id_cliente,
            c.nombre,
            c.apellido,
            c.ruc_cedula,
            c.email,
            c.telefono,
            c.direccion,
            'SI' if c.credito_activo else 'NO',
            c.limite_credito,
            c.dias_plazo
        ])

    wb.save(response)
    return response

@login_required
def importar_clientes_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                count_creados = 0
                count_actualizados = 0
                errores = []

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]: continue
                    
                    try:
                        id_cliente = str(row[0]).strip()
                        nombre = row[1]
                        apellido = row[2]
                        
                        defaults = {
                            'user': request.user, # Assign user
                            'nombre': nombre,
                            'apellido': apellido,
                            'ruc_cedula': row[3],
                            'email': row[4],
                            'telefono': row[5],
                            'direccion': row[6],
                            'credito_activo': True if row[7] == 'SI' else False,
                            'limite_credito': Decimal(str(row[8])) if row[8] is not None else 0,
                            'dias_plazo': int(row[9]) if row[9] is not None else 30
                        }
                        
                        # Note: This might fail if ID exists for another user (Global PK)
                        obj, created = Cliente.objects.update_or_create(
                            id_cliente=id_cliente,
                            defaults=defaults
                        )
                        
                        if created: count_creados += 1
                        else: count_actualizados += 1
                            
                    except Exception as e:
                        errores.append(f"Fila {row_idx}: {str(e)}")
                
                messages.success(request, f"Importación completada: {count_creados} creados, {count_actualizados} actualizados.")
                if errores:
                    messages.warning(request, f"Errores: {', '.join(errores[:3])}...")
            
            except Exception as e:
                 messages.error(request, f"Error: {str(e)}")
                 
            return redirect('clientes:lista')
    return redirect('clientes:lista')

@login_required
def clientes_lista(request):
    clientes = Cliente.objects.filter(user=request.user)
    context = {
        'clientes': clientes,
    }
    return render(request, 'clientes/lista.html', context)

@login_required
def cliente_crear(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.user = request.user
            cliente.save()
            messages.success(request, 'Cliente creado exitosamente')
            return redirect('clientes:lista')
        else:
            messages.error(request, 'Error al crear el cliente. Verifica los datos.')
    else:
        form = ClienteForm()
    
    context = {
        'form': form,
        'accion': 'Crear'
    }
    return render(request, 'clientes/form.html', context)

@login_required
def cliente_editar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente actualizado exitosamente')
            return redirect('clientes:lista')
        else:
            messages.error(request, 'Error al actualizar el cliente.')
    else:
        form = ClienteForm(instance=cliente)
    
    context = {
        'form': form,
        'cliente': cliente,
        'accion': 'Editar'
    }
    return render(request, 'clientes/form.html', context)

@login_required
def cliente_eliminar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, user=request.user)
    
    if request.method == 'POST':
        nombre = cliente.nombre_completo()
        cliente.delete()
        messages.success(request, f'Cliente {nombre} eliminado exitosamente')
        return redirect('clientes:lista')
    
    return redirect('clientes:lista')

@login_required
def cliente_detalle(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, user=request.user)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from possitema.credit_utils import obtener_resumen_credito_cliente
        
        # Información básica
        data = {
            'id_cliente': cliente.id_cliente,
            'nombre': cliente.nombre,
            'apellido': cliente.apellido,
            'ruc_cedula': cliente.ruc_cedula,
            'email': cliente.email,
            'telefono': cliente.telefono,
            'direccion': cliente.direccion,
            'fecha_registro': cliente.fecha_registro.strftime('%d/%m/%Y %H:%M'),
        }
        
        # Agregar información de crédito
        if cliente.credito_activo:
            resumen = obtener_resumen_credito_cliente(cliente)
            data['credito'] = {
                'activo': cliente.credito_activo,
                'limite': float(cliente.limite_credito),
                'disponible': float(cliente.saldo_credito),
                'usado': float(cliente.limite_credito - cliente.saldo_credito),
                'dias_plazo': cliente.dias_plazo,
                'cuentas_activas': resumen['cuentas_activas'],
                'saldo_total_pendiente': float(resumen['saldo_total_pendiente']),
            }
        else:
            data['credito'] = {
                'activo': False,
                'limite': 0,
                'disponible': 0,
                'usado': 0,
                'dias_plazo': 0,
                'cuentas_activas': 0,
                'saldo_total_pendiente': 0,
            }
        
        return JsonResponse(data)
    
    context = {
        'cliente': cliente,
    }
    return render(request, 'clientes/detalle.html', context)

@login_required
def cliente_actualizar_ruc(request, pk):
    """
    Endpoint AJAX para actualizar rápidamente el RUC/Cédula de un cliente
    """
    if request.method == 'POST':
        cliente = get_object_or_404(Cliente, pk=pk, user=request.user)
        import json
        
        try:
            data = json.loads(request.body)
            ruc_cedula = data.get('ruc_cedula', '').strip()
            
            if ruc_cedula:
                # Verificar que el RUC no esté duplicado
                if Cliente.objects.filter(user=request.user, ruc_cedula=ruc_cedula).exclude(id_cliente=pk).exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Este RUC/Cédula ya está registrado para otro cliente.'
                    }, status=400)
                
                cliente.ruc_cedula = ruc_cedula
                cliente.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'RUC/Cédula actualizado: {ruc_cedula}'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'El RUC/Cédula no puede estar vacío.'
                }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def buscar_clientes_api(request):
    """API para buscar clientes por nombre, email o teléfono"""
    from django.http import JsonResponse
    from django.db.models import Q
    
    try:
        q = request.GET.get('q', '').strip()
        
        # Permitir búsquedas con 1 carácter
        if not q or len(q) < 1:
            return JsonResponse({'clientes': []})
        
        # Buscar clientes - ser más flexible con la búsqueda
        clientes = Cliente.objects.filter(
            user=request.user,
        ).filter(
            Q(nombre__icontains=q) |
            Q(email__icontains=q) |
            Q(telefono__icontains=q) |
            Q(id_cliente__icontains=q)  # También permitir búsqueda por ID
        ).order_by('nombre').values('id_cliente', 'nombre', 'email', 'telefono')[:15]
        
        # Convertir id_cliente a id para que el frontend sea compatible
        clientes_list = []
        for cliente in clientes:
            clientes_list.append({
                'id': str(cliente['id_cliente']),
                'nombre': cliente['nombre'],
                'email': cliente['email'] or '',
                'telefono': cliente['telefono'] or ''
            })
        
        return JsonResponse({'clientes': clientes_list})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

