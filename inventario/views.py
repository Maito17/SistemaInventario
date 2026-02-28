from django.shortcuts import render
# inventario/views.py
from django.http import JsonResponse
from django.db.models import Q, F
from django.contrib.auth.decorators import login_required, permission_required
from .models import Producto, Proveedor, Categoria, Compra, DetalleCompra
from .forms import ProductoForm, ProveedorForm, CategoriaForm, ExcelUploadForm
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Count
import json 
import openpyxl
import logging
from io import BytesIO
from base64 import b64encode
import qrcode 
from django.http import HttpResponse
from django.contrib import messages

from datetime import timedelta, date
from decimal import Decimal
from finanzas.models import CuentaPorPagar

logger = logging.getLogger(__name__)

BAJO_STOCK_UMBRAL = 5
DIAS_ALERTA_CADUCIDAD = 30  # Alerta cuando falta 30 días o menos para vencer


def _obtener_metodos_pago_compra():
    """Retorna métodos de pago de compra; si no existen, crea los predeterminados."""
    from .models import MetodoPagoCompra

    metodos_pago = MetodoPagoCompra.objects.all()
    if metodos_pago.exists():
        return metodos_pago

    codigos_default = [
        'EFECTIVO',
        'TRANSFERENCIA',
        'CHEQUE',
        'CREDITO',
        'TARJETA',
        'OTRO',
    ]
    for codigo in codigos_default:
        MetodoPagoCompra.objects.get_or_create(nombre=codigo)

    return MetodoPagoCompra.objects.all()

@login_required
def buscar_producto_ajax(request):
    """
    Endpoint API para buscar productos por nombre o SKU/código en tiempo real.
    Devuelve resultados en formato JSON.
    
    CORRECCIÓN: Se elimina la lógica de parseo JSON ya que los QR ahora solo codifican el ID (pk).
    """
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Obtener el término de búsqueda de la solicitud GET
        query = request.GET.get('q', '')
        
        if query:
            # 🛑 1. ANTES: Se intentaba parsear JSON. AHORA: Simplemente usamos la query como ID/nombre.
            #     Esto soporta la lectura directa del QR (solo el ID: "BTP001").
            
            # 2. Buscar productos usando la 'query'
            # Se busca por nombre o el Primary Key (pk) del producto.
            productos = Producto.objects.filter(
                user=request.user,
                nombre__icontains=query,
                cantidad__gt=0
            ).values('id_producto', 'nombre', 'precio_venta', 'cantidad')[:10]
            
            # Convertir el QuerySet a una lista para el JSON
            lista_productos = list(productos)
            
            return JsonResponse({'productos': lista_productos})
        
        # Si la query está vacía, devuelve una lista vacía
        return JsonResponse({'productos': []})
    
    # Si no es una solicitud AJAX, deniega el acceso
    return JsonResponse({'error': 'Acceso denegado'}, status=400)


@login_required
def buscar_productos_api(request):
    """API para buscar productos para solicitud de crédito"""
    from django.http import JsonResponse
    from django.db.models import Q
    
    try:
        q = request.GET.get('q', '').strip()
        
        if not q or len(q) < 1:
            return JsonResponse({'productos': []})
        
        # Buscar productos
        from django.db.models import Q
        productos = Producto.objects.filter(
            Q(nombre__icontains=q) | Q(id_producto__icontains=q),
            cantidad__gt=0
        ).order_by('nombre').values('id_producto', 'nombre', 'precio_venta', 'cantidad')[:15]
        
        # Convertir a lista de diccionarios
        productos_list = []
        for p in productos:
            productos_list.append({
                'id': str(p['id_producto']),
                'nombre': p['nombre'],
                'precio': float(p['precio_venta']),
                'cantidad_disponible': int(p['cantidad'])
            })
        
        return JsonResponse({'productos': productos_list})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@permission_required('inventario.view_producto', raise_exception=True)
def productos_lista(request):
    """Muestra la lista de productos."""
    productos = Producto.objects.all()
    
    # Motor de búsqueda
    query = request.GET.get('q', '').strip()
    categoria_filtro = request.GET.get('categoria', '').strip()
    stock_filtro = request.GET.get('stock', '').strip()
    
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(id_producto__icontains=query) |
            Q(descripcion__icontains=query)
        )
    
    if categoria_filtro:
        productos = productos.filter(categoria__id=categoria_filtro)
    
    if stock_filtro == 'bajo':
        productos = productos.filter(cantidad__lte=BAJO_STOCK_UMBRAL)
    elif stock_filtro == 'sin':
        productos = productos.filter(cantidad=0)
    elif stock_filtro == 'disponible':
        productos = productos.filter(cantidad__gt=0)
    
    # Obtener categorías para el filtro
    categorias = Categoria.objects.all()
    
    # 2. Calcular métricas para los cuadros (Info Boxes)
    total_productos = Producto.objects.count()
    
    # Opcional: Calcular el valor total del inventario (si tienes precio_costo o precio_venta)
    # Asumiendo que el campo de stock es 'cantidad' y el de precio es 'precio_venta'
    valor_inventario = productos.aggregate(total=Sum('precio_venta', default=0))['total']
    
    # Calcular el total de unidades en stock
    total_stock = productos.aggregate(total=Sum('cantidad', default=0))['total']

    # Contar productos con bajo stock
    productos_bajo_stock = productos.filter(cantidad__lte=BAJO_STOCK_UMBRAL).count()
    
    # Contar productos próximos a vencer
    productos_proximos_a_vencer = get_productos_por_vencer(None).count()  # Ajusta la función para no filtrar por usuario
    
    context = {
        'productos': productos,
        # Métricas para los cuadros
        'total_productos': total_productos,
        'valor_inventario': valor_inventario,
        'total_stock': total_stock,
        'productos_bajo_stock': productos_bajo_stock,
        'productos_proximos_a_vencer': productos_proximos_a_vencer,
        # Búsqueda y filtros
        'query': query,
        'categorias': categorias,
        'categoria_filtro': categoria_filtro,
        'stock_filtro': stock_filtro,
    }
    # Asegúrate de que tienes un template en inventario/templates/inventario/lista_productos.html
    return render(request, 'inventario/lista_productos.html', context)

@login_required
@permission_required('inventario.add_producto', raise_exception=True)
def producto_crear(request):
    """Permite crear un nuevo producto, respetando el límite del plan."""
    # Eliminada la verificación de límite de productos por plan
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.user = request.user
            producto.save()
            return redirect('inventario:lista')
    else:
        form = ProductoForm()
    return render(request, 'inventario/crear_producto.html', {'form': form})

@login_required
@permission_required('inventario.add_proveedor', raise_exception=True)
def proveedor_crear(request):
    """Permite crear un nuevo proveedor."""
    if request.method == 'POST':
        form = ProveedorForm(request.POST) # 🛑 Usar ProveedorForm
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.user = request.user
            proveedor.save()
            # Redirigir a la lista de proveedores después de guardar
            return redirect('inventario:proveedores_lista') 
    else:
        form = ProveedorForm() # 🛑 Pasar ProveedorForm
        
    return render(request, 'inventario/crear_proveedor.html', {'form': form})


@login_required
@permission_required('inventario.add_categoria', raise_exception=True)
def categoria_crear(request):
    """Permite crear una nueva categoría."""
    if request.method == 'POST':
        form = CategoriaForm(request.POST) # 🛑 Usar CategoriaForm
        if form.is_valid():
            categoria = form.save(commit=False)
            categoria.user = request.user
            categoria.save()
            # Redirigir a la lista de categorías después de guardar
            return redirect('inventario:categorias_lista') 
    else:
        form = CategoriaForm() # 🛑 Pasar CategoriaForm
        
    return render(request, 'inventario/crear_categoria.html', {'form': form})

@login_required
@permission_required('inventario.view_proveedor', raise_exception=True)
def proveedores_lista(request): # 🛑 Asegúrate de que el nombre coincida EXACTAMENTE
    """Muestra la lista de proveedores."""
    proveedores = Proveedor.objects.all()
    
    # Motor de búsqueda
    query = request.GET.get('q', '').strip()
    
    if query:
        proveedores = proveedores.filter(
            Q(nombre__icontains=query) |
            Q(id_proveedor__icontains=query) |
            Q(contacto__icontains=query) |
            Q(telefono__icontains=query) |
            Q(email__icontains=query)
        )
    
    context = {
        'proveedores': proveedores,
        'query': query,
    }
    # Renderizamos la plantilla que creamos antes
    return render(request, 'inventario/lista_proveedores.html', context)

@login_required
@permission_required('inventario.view_categoria', raise_exception=True)
def categorias_lista(request): # 🛑 Asegúrate de que el nombre coincida EXACTAMENTE
    """Muestra la lista de categorías."""
    categorias = Categoria.objects.all()
    return render(request, 'inventario/lista_categorias.html', {'categorias': categorias})


# editar_producto definido más abajo con soporte AJAX/modal

@login_required
@permission_required('inventario.delete_producto', raise_exception=True)
def eliminar_producto(request, pk):
    """Permite eliminar un producto."""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.delete()
        return redirect('inventario:lista')
    return render(request, 'inventario/eliminar_producto.html', {'producto': producto})

@login_required
@permission_required('inventario.view_producto', raise_exception=True)
def detalle_producto(request, pk):
    """Muestra los detalles de un producto."""
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'inventario/detalle_producto.html', {'producto': producto})


# 🛑 ÚNICA DEFINICIÓN DE LA VISTA QR 
@login_required
def generar_qr_producto(request, pk):
    """
    Genera la página de impresión con los datos del producto y su código QR.
    El código QR contendrá SÓLO el ID/PK del producto para simplificar la lectura.
    """
    producto = get_object_or_404(Producto, pk=pk)

    # 🛑 Solamente se codifica el Primary Key (PK) del producto.
    qr_string = str(producto.pk) 

    # Generar el código QR como imagen (PNG en base64)
    qr_img = qrcode.make(qr_string, box_size=5, border=2)
    
    # Guardar la imagen en un buffer y codificarla a base64
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    qr_base64 = b64encode(buffer.getvalue()).decode()

    # Contexto a enviar a la plantilla
    context = {
        'producto': producto,
        'qr_base64_data': qr_base64,
        'qr_contenido_string': qr_string # Esto mostrará el PK directamente debajo del QR
    }

    return render(request, 'inventario/imprimir_qr_producto.html', context)

@login_required
@permission_required('inventario.change_producto', raise_exception=True)
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    
    # Manejar el envío del formulario
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            # Si es AJAX (desde la modal), solo devolvemos una respuesta de éxito 200
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return HttpResponse(status=200)
            # Si no es AJAX, redirigimos como antes
            return redirect('inventario:lista')
    else:
        # Mostrar el formulario
        form = ProductoForm(instance=producto)

    # Lógica de renderizado
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # SOLICITUD AJAX: Renderizar solo el formulario para la modal
        # NOTA: Debes crear un template MÁS PEQUEÑO que solo contenga el formulario: 'inventario/solo_formulario_edicion.html'
        # Pero, por simplicidad, usaremos el mismo template y esperaremos que el JS lo extraiga (menos eficiente)
        # Una mejor práctica es crear un nuevo template:
        return render(request, 'inventario/editar_producto_modal_content.html', {'form': form, 'producto': producto})
        
    else:
        # SOLICITUD NORMAL: Renderizar la página completa (no debería ocurrir con la modal, pero es un fallback)
        return render(request, 'inventario/editar_producto.html', {'form': form, 'producto': producto})

@login_required
@permission_required('inventario.change_proveedor', raise_exception=True)
def editar_proveedor(request, pk):
    """Permite editar un proveedor existente."""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('inventario:proveedores_lista')  # 🛑 CORRECCIÓN AQUÍ
    else:
        form = ProveedorForm(instance=proveedor)
    
    return render(request, 'inventario/editar_proveedor.html', {'form': form, 'proveedor': proveedor})

@login_required
@permission_required('inventario.delete_proveedor', raise_exception=True)
def eliminar_proveedor(request, pk):
    """Permite eliminar un proveedor."""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        proveedor.delete()
        return redirect('inventario:proveedores_lista')  # 🛑 CORRECCIÓN AQUÍ
    
    return render(request, 'inventario/eliminar_proveedor.html', {'proveedor': proveedor})

@login_required
@permission_required('inventario.change_categoria', raise_exception=True)
def editar_categoria(request, pk):
    """Permite editar una categoría existente."""
    categoria = get_object_or_404(Categoria, pk=pk)
    
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            return redirect('inventario:categorias_lista')
    else:
        form = CategoriaForm(instance=categoria)
    
    return render(request, 'inventario/editar_categoria.html', {'form': form, 'categoria': categoria})

@login_required
@permission_required('inventario.delete_categoria', raise_exception=True)
def eliminar_categoria(request, pk):
    """Permite eliminar una categoría."""
    categoria = get_object_or_404(Categoria, pk=pk)
    
    if request.method == 'POST':
        categoria.delete()
        return redirect('inventario:categorias_lista')
    
    return render(request, 'inventario/eliminar_categoria.html', {'categoria': categoria})

@login_required
@permission_required('inventario.view_categoria', raise_exception=True)
def detalle_categoria(request, pk):
    """Muestra los detalles de una categoría."""
    categoria = get_object_or_404(Categoria, pk=pk)
    return render(request, 'inventario/detalle_categoria.html', {'categoria': categoria})

def check_categoria_caducidad(request, pk):
    """Vista AJAX que verifica si una categoría requiere fecha de caducidad."""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.method == 'GET':
        try:
            categoria = Categoria.objects.get(pk=pk)
            data = {
                'requiere_caducidad': categoria.requiere_caducidad
            }
            return JsonResponse(data)
        except Categoria.DoesNotExist:
            return JsonResponse({'error': 'Categoría no encontrada'}, status=404)
    
    return JsonResponse({'error': 'Método no permitido'}, status=400)

#notficacion de producto por vencer 
def get_productos_por_vencer(user, dias_alerta=None):
    if dias_alerta is None:
        dias_alerta = DIAS_ALERTA_CADUCIDAD
    fecha_limite = date.today()+timedelta(days=dias_alerta)
    productos_lista = Producto.objects.filter(
        user=user,
        fecha_caducidad__isnull=False,
        fecha_caducidad__lte=fecha_limite,
        
    ).exclude(
        cantidad=0
    ).order_by('fecha_caducidad')
    return productos_lista
@login_required
def dashboard_view(request):
    productos_vencidos = get_productos_por_vencer(request.user, dias_alerta=0)
    conteo_vencimientos = productos_vencidos.count()

    productos_bajo_stock_count = Producto.objects.filter(
        user=request.user,
        cantidad__lte=BAJO_STOCK_UMBRAL
    ).count()
    
    context = {
        'conteo_vencimientos': conteo_vencimientos,
        'productos_bajo_stock': productos_bajo_stock_count,
    }
    return render(request, 'base.html', context)

@login_required
def get_alertas_caducidad_ajax(request):
    from datetime import date
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.method == 'GET':
        dias_alerta = int(request.GET.get('dias_alerta', 30))
        productos_por_vencer = get_productos_por_vencer(request.user, dias_alerta)
        
        alertas = []
        for producto in productos_por_vencer:
            if producto.fecha_caducidad:
                dias_restantes = (producto.fecha_caducidad - date.today()).days
                alertas.append({
                    'id': producto.id_producto,
                    'nombre': producto.nombre,
                    'fecha_caducidad': str(producto.fecha_caducidad),
                    'cantidad': producto.cantidad,
                    'dias_restantes': dias_restantes
                })
        
        return JsonResponse({
            'count': len(alertas),
            'alertas': alertas
        })

@login_required
@permission_required('inventario.view_compra', raise_exception=True)
def get_alertas_compras_ajax(request):
    """Obtiene las compras pendientes de pago a través de AJAX."""
    from datetime import date, timedelta
    from .models import Compra
    
    try:
        # Obtener compras que tengan fecha de pago proveedor asignada y NO estén en estado PAGADA o CANCELADA
        compras_pendientes = Compra.objects.filter(
            user=request.user,
            fecha_pago_proveedor__isnull=False
        ).exclude(
            estado__in=['PAGADA', 'CANCELADA']
        ).order_by('fecha_pago_proveedor')
        
        alertas = []
        hoy = date.today()
        
        for compra in compras_pendientes:
            dias_restantes = (compra.fecha_pago_proveedor - hoy).days
            
            # Mostrar todas las compras con fecha de pago, sin importar los días
            alertas.append({
                'id': compra.id_compra,
                'numero': compra.numero_documento or f'Compra #{compra.id_compra}',
                'proveedor': compra.proveedor.nombre if compra.proveedor else 'Sin proveedor',
                'fecha_pago': str(compra.fecha_pago_proveedor),
                'dias_restantes': dias_restantes,
                'monto': str(compra.total),
                'url_editar': f'/inventario/compras/{compra.id_compra}/editar/'
            })
        
        return JsonResponse({
            'count': len(alertas),
            'alertas': alertas
        })
    except Exception as e:
        import traceback
        logger.error('Error en get_alertas_compras_ajax: %s', str(e), exc_info=True)
        return JsonResponse({
            'count': 0,
            'alertas': [],
            'error': str(e)
        }, status=500)


def get_alertas_bajo_stock_ajax(request):
    """Obtiene los productos con bajo stock a través de AJAX."""
    from .models import Producto
    
    try:
        logger.debug('get_alertas_bajo_stock_ajax llamado')
        
        # Obtener productos con cantidad <= BAJO_STOCK_UMBRAL (definido al inicio del archivo)
        # El umbral es consistente con el dashboard (5 unidades)
        UMBRAL_BAJO_STOCK = BAJO_STOCK_UMBRAL  # 5 unidades
        productos_bajo_stock = Producto.objects.filter(
            estado='ACTIVO',
            cantidad__lte=UMBRAL_BAJO_STOCK
        ).order_by('cantidad')
        
        logger.debug('Productos encontrados (cantidad <= %d): %d', UMBRAL_BAJO_STOCK, productos_bajo_stock.count())
        
        productos_list = []
        for producto in productos_bajo_stock:
            productos_list.append({
                'id': producto.id_producto,
                'nombre': producto.nombre,
                'sku': producto.id_producto,
                'cantidad': producto.cantidad
            })
        
        response_data = {
            'count': len(productos_list),
            'productos': productos_list
        }
        logger.debug('Respuesta count: %d', response_data['count'])
        
        return JsonResponse(response_data)
    except Exception as e:
        import traceback
        error_msg = f"Error en get_alertas_bajo_stock_ajax: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return JsonResponse({
            'count': 0,
            'productos': [],
            'error': error_msg
        }, status=500)




# --- VISTAS PARA COMPRAS Y DETALLES DE COMPRAS ---

@login_required
@permission_required('inventario.view_compra', raise_exception=True)
def lista_compras(request):
    """Muestra el listado de todas las compras registradas."""
    from django.core.paginator import Paginator
    from .models import Compra
    
    compras_list = Compra.objects.all().order_by('-fecha_compra')
    
    # Motor de búsqueda
    query = request.GET.get('q', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    
    if query:
        compras_list = compras_list.filter(
            Q(id_compra__icontains=query) |
            Q(proveedor__nombre__icontains=query) |
            Q(numero_documento__icontains=query)
        )
    
    if estado_filtro:
        compras_list = compras_list.filter(estado=estado_filtro)
    
    if fecha_desde:
        compras_list = compras_list.filter(fecha_compra__date__gte=fecha_desde)
    
    if fecha_hasta:
        compras_list = compras_list.filter(fecha_compra__date__lte=fecha_hasta)
    
    # Paginación
    paginator = Paginator(compras_list, 10)
    page_number = request.GET.get('page')
    compras = paginator.get_page(page_number)
    
    # Query string sin 'page' para paginación
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()
    
    context = {
        'compras': compras,
        'paginator': paginator,
        'query': query,
        'estado_filtro': estado_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'query_string': query_string,
    }
    return render(request, 'inventario/lista_compras.html', context)


@login_required
@permission_required('inventario.add_compra', raise_exception=True)
def crear_compra(request):
    """Vista para crear una nueva compra con detalles inline."""
    from .models import Compra, DetalleCompra, Producto
    from django.db import transaction
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Crear la compra
                proveedor_id = request.POST.get('proveedor')
                numero_documento = request.POST.get('numero_documento', '')
                metodo_pago_id = request.POST.get('metodo_pago')
                estado = request.POST.get('estado', 'PENDIENTE')
                fecha_pago_proveedor = request.POST.get('fecha_pago_proveedor', '')
                
                compra = Compra.objects.create(
                    user=request.user,
                    proveedor_id=proveedor_id,
                    numero_documento=numero_documento,
                    metodo_pago_id=metodo_pago_id if metodo_pago_id else None,
                    estado=estado,
                    fecha_pago_proveedor=fecha_pago_proveedor if fecha_pago_proveedor else None,
                )
                
                # Procesar los detalles de la compra
                cantidad_detalles = int(request.POST.get('detalles-TOTAL_FORMS', 0))
                total_compra = 0
                detalles_guardados = 0
                
                for i in range(cantidad_detalles):
                    prefix = f'detalles-{i}'
                    producto_id = request.POST.get(f'{prefix}-producto')
                    cantidad = request.POST.get(f'{prefix}-cantidad_recibida')
                    costo_unitario = request.POST.get(f'{prefix}-costo_unitario')
                    fecha_caducidad = request.POST.get(f'{prefix}-fecha_caducidad')
                    metodo_pago_id = request.POST.get(f'{prefix}-metodo_pago')
                    
                    if not producto_id and not cantidad and not costo_unitario:
                        continue

                    if not producto_id or not cantidad or not costo_unitario:
                        raise ValueError(f'Línea {i + 1}: producto, cantidad y costo son obligatorios.')

                    try:
                        cantidad = int(str(cantidad).strip())
                        costo_unitario = float(str(costo_unitario).strip().replace(',', '.'))

                        if cantidad <= 0:
                            raise ValueError(f'Línea {i + 1}: la cantidad debe ser mayor a 0.')

                        if costo_unitario <= 0:
                            raise ValueError(f'Línea {i + 1}: el costo unitario debe ser mayor a 0.')

                        detalle = DetalleCompra.objects.create(
                            compra=compra,
                            producto_id=producto_id,
                            cantidad_recibida=cantidad,
                            costo_unitario=costo_unitario,
                            fecha_caducidad=fecha_caducidad if fecha_caducidad else None,
                            metodo_pago_id=metodo_pago_id if metodo_pago_id else None,
                        )

                        total_compra += detalle.subtotal
                        detalles_guardados += 1
                    except (ValueError, TypeError) as e:
                        logger.warning('Error al procesar detalle: %s', str(e))
                        raise ValueError(str(e))

                if detalles_guardados == 0:
                    raise ValueError('Debes agregar al menos un producto válido con cantidad y costo unitario.')
                
                # Actualizar el total de la compra
                compra.total = total_compra
                compra.save()
                
                # Si la compra se crea como RECIBIDA, crear la CuentaPorPagar
                if compra.estado == 'RECIBIDA':
                    logger.debug('Intentando crear CuentaPorPagar para compra %s (usuario: %s)', compra.id_compra, request.user)
                    try:
                        cuenta = CuentaPorPagar.objects.create(
                            compra=compra,
                            monto_total=compra.total,
                            saldo=compra.total,
                            fecha_vencimiento=compra.fecha_pago_proveedor,
                            estado='PENDIENTE',
                            owner=request.user
                        )
                        logger.debug('CuentaPorPagar creada: %s', cuenta)
                    except Exception as e:
                        logger.error('No se pudo crear CuentaPorPagar: %s', e)
                
                messages.success(request, f'Compra #{compra.id_compra} creada exitosamente.')
                return redirect('inventario:lista_compras')
        except Exception as e:
            messages.error(request, f'Error al crear la compra: {str(e)}')
    
    from .models import Proveedor, MetodoPagoCompra
    
    proveedores = Proveedor.objects.all()
    productos = Producto.objects.all()  # Mostrar todos los productos, incluso los agotados
    metodos_pago = _obtener_metodos_pago_compra()
    
    context = {
        'proveedores': proveedores,
        'productos': productos,
        'metodos_pago': metodos_pago,
    }
    return render(request, 'inventario/crear_compra.html', context)


@login_required
@permission_required('inventario.view_detallecompra', raise_exception=True)
def lista_detalles_compra(request):
    """Muestra el listado de todos los detalles de compra."""
    from django.core.paginator import Paginator
    from .models import DetalleCompra
    
    detalles_list = DetalleCompra.objects.select_related('compra', 'producto').order_by('-compra__fecha_compra')
    
    # Motor de búsqueda
    query = request.GET.get('q', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    
    if query:
        detalles_list = detalles_list.filter(
            Q(compra__id_compra__icontains=query) |
            Q(producto__nombre__icontains=query) |
            Q(producto__id_producto__icontains=query) |
            Q(compra__proveedor__nombre__icontains=query)
        )
    
    if fecha_desde:
        detalles_list = detalles_list.filter(compra__fecha_compra__date__gte=fecha_desde)
    
    if fecha_hasta:
        detalles_list = detalles_list.filter(compra__fecha_compra__date__lte=fecha_hasta)
    
    # Paginación
    paginator = Paginator(detalles_list, 15)
    page_number = request.GET.get('page')
    detalles = paginator.get_page(page_number)
    
    # Query string sin 'page' para paginación
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()
    
    context = {
        'detalles': detalles,
        'paginator': paginator,
        'query': query,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'query_string': query_string,
    }
    return render(request, 'inventario/lista_detalles_compra.html', context)


@login_required
@permission_required('inventario.change_compra', raise_exception=True)
def editar_compra(request, pk):
    """Vista para editar una compra existente."""
    from .models import Compra, DetalleCompra, Producto, MetodoPagoCompra
    from django.db import transaction
    from decimal import Decimal
    
    compra = get_object_or_404(Compra, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Actualizar los datos generales de la compra
                proveedor_id = request.POST.get('proveedor')
                numero_documento = request.POST.get('numero_documento', '')
                metodo_pago_id = request.POST.get('metodo_pago')
                estado = request.POST.get('estado', 'PENDIENTE')
                fecha_pago_proveedor = request.POST.get('fecha_pago_proveedor', '')
                
                if not proveedor_id:
                    messages.error(request, 'El proveedor es requerido.')
                    return redirect('inventario:editar_compra', pk=pk)
                
                compra.proveedor_id = proveedor_id
                compra.numero_documento = numero_documento
                compra.metodo_pago_id = metodo_pago_id if metodo_pago_id else None
                compra.estado = estado
                compra.fecha_pago_proveedor = fecha_pago_proveedor if fecha_pago_proveedor else None
                
                # Si el estado cambia a RECIBIDA, guardar la fecha de recepción
                if estado == 'RECIBIDA' and not compra.fecha_recibida:
                    from django.utils import timezone
                    compra.fecha_recibida = timezone.now()
                
                # Eliminar detalles existentes (para recriarlos)
                compra.detalles.all().delete()
                
                # Procesar los detalles de la compra desde el formulario
                cantidad_detalles = int(request.POST.get('detalles-TOTAL_FORMS', 0))
                total_compra = Decimal('0')
                detalles_guardados = 0
                
                for i in range(cantidad_detalles):
                    prefix = f'detalles-{i}'
                    
                    # Obtener datos del formulario
                    producto_id = request.POST.get(f'{prefix}-producto', '').strip()
                    cantidad_str = request.POST.get(f'{prefix}-cantidad_recibida', '').strip()
                    costo_str = request.POST.get(f'{prefix}-costo_unitario', '').strip()
                    fecha_caducidad = request.POST.get(f'{prefix}-fecha_caducidad', '')
                    metodo_pago_id_detalle = request.POST.get(f'{prefix}-metodo_pago')
                    
                    # Validar que los campos requeridos estén presentes
                    if not producto_id or not cantidad_str or not costo_str:
                        continue
                    
                    try:
                        cantidad = int(cantidad_str)
                        costo_unitario = Decimal(costo_str.replace(',', '.'))
                        
                        # Validaciones
                        if cantidad <= 0:
                            messages.warning(request, f'Línea {i+1}: La cantidad debe ser mayor a 0.')
                            continue
                        
                        if costo_unitario < 0:
                            messages.warning(request, f'Línea {i+1}: El costo no puede ser negativo.')
                            continue
                        
                        # Verificar que el producto existe
                        producto = Producto.objects.get(id_producto=producto_id)
                        
                        # Crear el detalle
                        detalle = DetalleCompra.objects.create(
                            compra=compra,
                            producto=producto,
                            cantidad_recibida=cantidad,
                            costo_unitario=costo_unitario,
                            fecha_caducidad=fecha_caducidad if fecha_caducidad else None,
                            metodo_pago_id=metodo_pago_id_detalle if metodo_pago_id_detalle else None,
                        )
                        
                        total_compra += detalle.subtotal
                        detalles_guardados += 1
                        
                    except (ValueError, TypeError) as e:
                        messages.warning(request, f'Línea {i+1}: Error en los datos - {str(e)}')
                        continue
                    except Producto.DoesNotExist:
                        messages.warning(request, f'Línea {i+1}: Producto no encontrado.')
                        continue
                
                # Validar que hay al menos un detalle
                if detalles_guardados == 0:
                    messages.error(request, 'Debe agregar al menos un detalle a la compra.')
                    return redirect('inventario:editar_compra', pk=pk)
                
                # Actualizar el total y guardar la compra
                compra.total = total_compra
                compra.save()
                
                # Si la compra cambia a RECIBIDA, crear o actualizar la CuentaPorPagar
                if compra.estado == 'RECIBIDA':
                    logger.debug('Intentando get_or_create CuentaPorPagar para compra %s (usuario: %s)', compra.id_compra, request.user)
                    try:
                        cuenta_por_pagar, created = CuentaPorPagar.objects.get_or_create(
                            compra=compra,
                            defaults={
                                'monto_total': compra.total,
                                'saldo': compra.total,
                                'fecha_vencimiento': compra.fecha_pago_proveedor,
                                'estado': 'PENDIENTE',
                                'owner': request.user
                            }
                        )
                        logger.debug('CuentaPorPagar %s: %s', 'creada' if created else 'encontrada', cuenta_por_pagar)
                        # Si ya existía, actualizar los montos y el owner
                        if not created:
                            cuenta_por_pagar.monto_total = compra.total
                            cuenta_por_pagar.saldo = compra.total - cuenta_por_pagar.monto_pagado
                            cuenta_por_pagar.fecha_vencimiento = compra.fecha_pago_proveedor
                            cuenta_por_pagar.owner = request.user
                            if cuenta_por_pagar.saldo <= 0:
                                cuenta_por_pagar.estado = 'PAGADA'
                            elif cuenta_por_pagar.monto_pagado > 0:
                                cuenta_por_pagar.estado = 'PARCIAL'
                            else:
                                cuenta_por_pagar.estado = 'PENDIENTE'
                            cuenta_por_pagar.save()
                    except Exception as e:
                        logger.error('No se pudo crear/actualizar CuentaPorPagar: %s', e)
                
                messages.success(request, f'✅ Compra #{compra.id_compra} actualizada exitosamente (Estado: {compra.get_estado_display()}) con {detalles_guardados} producto(s).')
                return redirect('inventario:lista_compras')
                
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar la compra: {str(e)}')
    
    from .models import Proveedor, MetodoPagoCompra
    
    proveedores = Proveedor.objects.all()
    productos = Producto.objects.all()  # Mostrar todos los productos, incluso los agotados
    metodos_pago = _obtener_metodos_pago_compra()
    detalles = compra.detalles.all()
    
    context = {
        'compra': compra,
        'proveedores': proveedores,
        'productos': productos,
        'metodos_pago': metodos_pago,
        'detalles': detalles,
        'es_edicion': True,
    }
    return render(request, 'inventario/editar_compra.html', context)


# ========== GENERADOR DE CÓDIGO QR PARA COMPRAS ==========
@login_required
@permission_required('inventario.view_compra', raise_exception=True)
def generar_qr_compra(request, pk):
    """Genera un PDF con código QR para cada producto de una compra."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime
    import io
    
    try:
        compra = Compra.objects.get(id_compra=pk)
        detalles = compra.detalles.all()
        
        # Crear buffer PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.3*inch, bottomMargin=0.3*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1f3864'),
            spaceAfter=8,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        label_style = ParagraphStyle(
            'Label',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            fontName='Helvetica-Bold'
        )
        
        # Título
        elements.append(Paragraph('ETIQUETAS DE PRODUCTOS - COMPRA #' + str(compra.id_compra), title_style))
        elements.append(Spacer(1, 0.15*inch))
        
        # Crear etiqueta para cada producto
        for idx, detalle in enumerate(detalles):
            # Datos para el QR: ID del producto + información
            qr_data = f"PROD:{detalle.producto.id_producto}|COMPRA:{compra.id_compra}|CANT:{detalle.cantidad_recibida}"
            
            # Generar código QR
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=8,
                border=1,
            )
            qr.add_data(qr_data)
            qr.make(fit=True)
            
            qr_img = qr.make_image(fill_color='black', back_color='white')
            
            # Convertir imagen a buffer
            qr_buffer = io.BytesIO()
            qr_img.save(qr_buffer, format='PNG')
            qr_buffer.seek(0)
            
            # Crear tabla para la etiqueta (QR a la izquierda, información a la derecha)
            fecha_cad = detalle.fecha_caducidad.strftime('%d/%m/%Y') if detalle.fecha_caducidad else 'SIN FECHA'
            
            label_data = [
                [
                    Image(qr_buffer, width=1.2*inch, height=1.2*inch),
                    [
                        Paragraph(f"<b>ID Producto:</b>", label_style),
                        Paragraph(f"{detalle.producto.id_producto}", styles['Heading2']),
                        Spacer(1, 0.05*inch),
                        Paragraph(f"<b>Nombre:</b>", label_style),
                        Paragraph(f"{detalle.producto.nombre}", styles['Normal']),
                        Spacer(1, 0.05*inch),
                        Paragraph(f"<b>Fecha Caducidad:</b>", label_style),
                        Paragraph(f"{fecha_cad}", ParagraphStyle(
                            'CaducidadStyle',
                            parent=styles['Normal'],
                            fontSize=12,
                            textColor=colors.HexColor('#d32f2f'),
                            fontName='Helvetica-Bold'
                        )),
                    ]
                ]
            ]
            
            label_table = Table(label_data, colWidths=[1.5*inch, 4*inch])
            label_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f5f5')),
                ('BORDER', (0, 0), (-1, -1), 2, colors.HexColor('#1f3864')),
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(label_table)
            
            # Agregar espacio entre etiquetas
            if idx < len(detalles) - 1:
                elements.append(Spacer(1, 0.2*inch))
            
            # Salto de página cada 2 etiquetas para mejor distribución
            if (idx + 1) % 2 == 0 and idx < len(detalles) - 1:
                elements.append(PageBreak())
        
        # Construir PDF
        doc.build(elements)
        
        # Retornar PDF
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="ETIQUETAS_COMPRA_{compra.id_compra}.pdf"'
        return response
        
    except Compra.DoesNotExist:
        return HttpResponse('Compra no encontrada', status=404)
    except Exception as e:
        return HttpResponse(f'Error generando etiquetas: {str(e)}', status=500)


@login_required
def exportar_productos_excel(request):
    """
    Exporta la lista de productos a un archivo Excel.
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_productos.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    headers = ['ID Producto', 'Nombre', 'Descripción', 'Precio Costo', 'Precio Venta', 'Stock', 'Categoría', 'Tarifa IVA', 'Estado', 'Fecha Caducidad']
    ws.append(headers)

    productos = Producto.objects.all()
    for p in productos:
        ws.append([
            p.id_producto,
            p.nombre,
            p.descripcion,
            p.precio_costo,
            p.precio_venta,
            p.cantidad,
            p.categoria.nombre if p.categoria else '',
            p.tarifa_iva, # Exportamos el código (ej: '15') para facilitar importación
            p.estado,     # Exportamos el código (ej: 'ACTIVO')
            p.fecha_caducidad
        ])

    wb.save(response)
    return response

@login_required
def importar_productos_excel(request):
    """
    Vista para importar productos desde un archivo Excel.
    Si el ID existe, actualiza; si no, crea.
    """
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                count_creados = 0
                count_actualizados = 0
                errores = []

                # Iterar filas (asumiendo que la fila 1 es encabezado)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]: # Si no hay ID, saltar
                        continue
                    
                    try:
                        id_prod = str(row[0]).strip()
                        nombre = row[1]
                        
                        # Datos por defecto / limpieza
                        descripcion = row[2] if row[2] else ''
                        precio_costo = Decimal(str(row[3])) if row[3] is not None else Decimal('0')
                        precio_venta = Decimal(str(row[4])) if row[4] is not None else Decimal('0')
                        cantidad = int(row[5]) if row[5] is not None else 0
                        categoria_nombre = row[6]
                        tarifa_iva = str(row[7]) if row[7] else '15'
                        estado = str(row[8]) if row[8] else 'ACTIVO'
                        fecha_caducidad = row[9] # openpyxl suele devolver datetime.date si la celda tiene formato fecha
                        
                        # Manejo de Categoría
                        categoria = None
                        if categoria_nombre:
                            categoria, _ = Categoria.objects.get_or_create(nombre=str(categoria_nombre).strip())

                        defaults = {
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'precio_costo': precio_costo,
                            'precio_venta': precio_venta,
                            'cantidad': cantidad,
                            'categoria': categoria,
                            'tarifa_iva': tarifa_iva,
                            'estado': estado,
                            'fecha_caducidad': fecha_caducidad if fecha_caducidad else None
                        }
                        
                        obj, created = Producto.objects.update_or_create(
                            id_producto=id_prod,
                            defaults=defaults
                        )
                        
                        if created:
                            count_creados += 1
                        else:
                            count_actualizados += 1
                            
                    except Exception as e:
                        errores.append(f"Fila {row_idx}: {str(e)}")
                
                messages.success(request, f"Importación completada: {count_creados} creados, {count_actualizados} actualizados.")
                if errores:
                    messages.warning(request, f"Errores en algunas filas: {', '.join(errores[:5])}...")
            
            except Exception as e:
                 messages.error(request, f"Error procesando el archivo: {str(e)}")
                 
            return redirect('inventario:lista')
    else:
        # Si no es POST, no mostramos nada, ya que lo haremos vía modal en la misma lista
        return redirect('inventario:lista')