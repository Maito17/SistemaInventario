# reportes/views.py
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse, FileResponse
import os
from pathlib import Path
from django.contrib.auth.decorators import login_required

# ==================== DESCARGA DE RESPALDOS ====================

from django.conf import settings
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def descargar_respaldo(request, filename):
    """Permite descargar un archivo de respaldo desde backup/respaldos."""
    respaldo_dir = Path(settings.BASE_DIR) / 'backup' / 'respaldos'
    archivo_path = respaldo_dir / filename
    if not archivo_path.exists() or not archivo_path.is_file():
        return HttpResponse('Archivo no encontrado.', status=404)
    response = FileResponse(open(archivo_path, 'rb'), as_attachment=True, filename=filename)
    return response

from .models import RespaldoArchivo

@login_required
def lista_respaldos(request):
    """Muestra los respaldos disponibles para descarga."""
    respaldos = RespaldoArchivo.objects.all().order_by('-fecha_creacion')
    context = {
        'respaldos': respaldos,
        'titulo': 'Respaldos disponibles',
    }
    return render(request, 'reportes/lista_respaldos.html', context)

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, F, Value, FloatField
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import date, datetime, timedelta
from decimal import Decimal
import logging

from inventario.models import Producto
from cliente.models import Cliente
from ventas.models import Venta, DetalleVenta
from control.models import RegistroAsistencia
from usuarios.models import EstadoCaja
from django.contrib.auth.models import User

logger = logging.getLogger(__name__)


@login_required
def inicio_reportes(request):
    """Vista principal de reportes"""
    context = {
        'titulo': 'Centro de Reportes',
    }
    return render(request, 'reportes/inicio.html', context)


# ==================== REPORTES DE ASISTENCIA ====================

@login_required
def reporte_asistencia(request):
    """Reporte de asistencia del personal"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    usuario_id = request.GET.get('usuario')
    
    # Valores por defecto
    hoy = date.today()
    if not fecha_inicio:
        fecha_inicio = (hoy - timedelta(days=7)).isoformat()
    if not fecha_fin:
        fecha_fin = hoy.isoformat()
    
    # Construir query
    registros = RegistroAsistencia.objects.all()
    
    if fecha_inicio:
        registros = registros.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha__lte=fecha_fin)
    if usuario_id:
        registros = registros.filter(user_id=usuario_id)
    
    registros = registros.order_by('-fecha', '-hora_entrada')
    
    # Estadísticas
    total_dias = registros.values('fecha').distinct().count()
    usuarios_activos = registros.values('user').distinct().count()
    
    context = {
        'titulo': 'Reporte de Asistencia',
        'registros': registros,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'usuario_id': usuario_id,
        'usuarios': User.objects.exclude(username='admin'),
        'total_dias': total_dias,
        'usuarios_activos': usuarios_activos,
    }
    return render(request, 'reportes/asistencia.html', context)


# ==================== REPORTES DE CAJA ====================

@login_required
def reporte_caja(request):
    """Reporte de apertura y cierre de caja"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    usuario_id = request.GET.get('usuario')
    estado = request.GET.get('estado')
    
    # Valores por defecto
    hoy = date.today()
    if not fecha_inicio:
        fecha_inicio = (hoy - timedelta(days=30)).isoformat()
    if not fecha_fin:
        fecha_fin = hoy.isoformat()
    
    # Construir query
    cajas = EstadoCaja.objects.all()
    
    if fecha_inicio:
        cajas = cajas.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        cajas = cajas.filter(fecha__lte=fecha_fin)
    if usuario_id:
        cajas = cajas.filter(user_id=usuario_id)
    if estado:
        cajas = cajas.filter(estado=estado)
    
    cajas = cajas.order_by('-fecha', '-timestamp_apertura')
    
    # Estadísticas
    total_abierta = cajas.filter(estado='ABIERTA').aggregate(
        total=Sum('monto_inicial')
    )['total'] or 0
    
    total_cerrada = cajas.filter(estado='CERRADA').aggregate(
        total=Sum('monto_cierre')
    )['total'] or 0
    
    total_diferencia = cajas.filter(estado='CERRADA').aggregate(
        diferencia=Sum(F('monto_cierre') - F('monto_inicial'), output_field=FloatField())
    )['diferencia'] or 0
    
    context = {
        'titulo': 'Reporte de Caja',
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'usuario_id': usuario_id,
        'estado': estado,
        'usuarios': User.objects.exclude(username='admin'),
        'total_abierta': total_abierta,
        'total_cerrada': total_cerrada,
        'total_diferencia': total_diferencia,
    }
    return render(request, 'reportes/caja.html', context)


# ==================== REPORTES DE VENTAS ====================

@login_required
def reporte_ventas(request):
    """Reporte de ventas por periodo"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_reporte = request.GET.get('tipo', 'diario')  # diario, producto, cliente
    
    # Valores por defecto
    hoy = date.today()
    if not fecha_inicio:
        fecha_inicio = (hoy - timedelta(days=30)).isoformat()
    if not fecha_fin:
        fecha_fin = hoy.isoformat()
    
    # Construir query base
    ventas = Venta.objects.filter(
        owner=request.user,
        fecha_venta__gte=fecha_inicio,
        fecha_venta__lte=fecha_fin
    )
    
    if tipo_reporte == 'producto':
        # Agrupar por producto
        ventas_agrupadas = DetalleVenta.objects.filter(
            venta__owner=request.user,
            venta__fecha_venta__gte=fecha_inicio,
            venta__fecha_venta__lte=fecha_fin
        ).values('producto__nombre').annotate(
            total_cantidad=Sum('cantidad'),
            total_monto=Sum(F('cantidad') * F('precio_unitario'), output_field=FloatField()),
            cantidad_ventas=Count('venta_id', distinct=True)
        ).order_by('-total_monto')
        
        context = {
            'titulo': 'Reporte de Ventas por Producto',
            'tipo_reporte': 'producto',
            'ventas': ventas_agrupadas,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
    
    elif tipo_reporte == 'cliente':
        # Agrupar por cliente
        ventas_agrupadas = Venta.objects.filter(
            owner=request.user,
            fecha_venta__gte=fecha_inicio,
            fecha_venta__lte=fecha_fin
        ).values('cliente__nombre').annotate(
            total_monto=Sum('total'),
            cantidad_compras=Count('id_venta')
        ).order_by('-total_monto')
        
        context = {
            'titulo': 'Reporte de Ventas por Cliente',
            'tipo_reporte': 'cliente',
            'ventas': ventas_agrupadas,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
    
    else:  # diario
        # Agrupar por día
        ventas_agrupadas = Venta.objects.filter(
            owner=request.user,
            fecha_venta__gte=fecha_inicio,
            fecha_venta__lte=fecha_fin
        ).extra(
            select={'fecha': 'DATE(fecha_venta)'}
        ).values('fecha').annotate(
            total_monto=Sum('total'),
            cantidad_ventas=Count('id_venta')
        ).order_by('-fecha')
        
        context = {
            'titulo': 'Reporte de Ventas Diario',
            'tipo_reporte': 'diario',
            'ventas': ventas_agrupadas,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
    
    # Totales generales
    total_ventas = ventas.aggregate(
        cantidad=Count('id_venta'),
        monto=Sum('total')
    )
    
    context['total_cantidad'] = total_ventas['cantidad'] or 0
    context['total_monto'] = total_ventas['monto'] or 0
    
    return render(request, 'reportes/ventas.html', context)


# ==================== REPORTES DE CLIENTES ====================

@login_required
def reporte_clientes(request):
    """Reporte de clientes y deudores"""
    from finanzas.models import CuentaPorCobrar
    
    tipo_reporte = request.GET.get('tipo', 'general')  # general, deudores, mejores
    
    if tipo_reporte == 'deudores':
        # Clientes que deben dinero
        clientes = Cliente.objects.filter(
            venta__cuenta_por_cobrar__estado__in=['PENDIENTE', 'PARCIAL'],
            venta__owner=request.user
        ).distinct().annotate(
            total_deuda=Sum(
                'venta__cuenta_por_cobrar__saldo',
                filter=Q(venta__cuenta_por_cobrar__estado__in=['PENDIENTE', 'PARCIAL']) & Q(venta__owner=request.user)
            ),
            cantidad_cuentas=Count('venta__cuenta_por_cobrar', distinct=True)
        ).order_by('-total_deuda')
        
        titulo = 'Reporte de Clientes Deudores'
    
    elif tipo_reporte == 'mejores':
        # Mejores clientes por monto de compra
        clientes = Cliente.objects.annotate(
            total_compras=Sum('venta__total', filter=Q(venta__owner=request.user)),
            cantidad_compras=Count('venta', filter=Q(venta__owner=request.user), distinct=True)
        ).exclude(
            total_compras__isnull=True
        ).order_by('-total_compras')[:20]
        
        titulo = 'Reporte de Mejores Clientes'
    
    else:  # general
        clientes = Cliente.objects.all().annotate(
            total_compras=Coalesce(Sum('venta__total', filter=Q(venta__owner=request.user)), 0),
            cantidad_compras=Count('venta', filter=Q(venta__owner=request.user), distinct=True),
            total_deuda=Coalesce(
                Sum(
                    'venta__cuenta_por_cobrar__saldo',
                    filter=Q(venta__cuenta_por_cobrar__estado__in=['PENDIENTE', 'PARCIAL']) & Q(venta__owner=request.user)
                ),
                0
            )
        ).order_by('-total_compras')
        
        titulo = 'Reporte General de Clientes'
    
    context = {
        'titulo': titulo,
        'clientes': clientes,
        'tipo_reporte': tipo_reporte,
    }
    return render(request, 'reportes/clientes.html', context)


# ==================== REPORTES DE INVENTARIO ====================

@login_required
def reporte_inventario(request):
    """Reporte de inventario y stock"""
    tipo_reporte = request.GET.get('tipo', 'general')  # general, bajo_stock
    
    if tipo_reporte == 'bajo_stock':
        # Productos con bajo stock
        productos = Producto.objects.filter(user=request.user, cantidad__lte=10).order_by('cantidad')
        titulo = 'Reporte de Productos con Bajo Stock'
    else:
        # Todos los productos
        productos = Producto.objects.filter(user=request.user).order_by('-cantidad')
        titulo = 'Reporte General de Inventario'
    
    # Calcular estadísticas
    total_productos = productos.count()
    bajo_stock = Producto.objects.filter(user=request.user, cantidad__lte=10).count()
    sin_stock = Producto.objects.filter(user=request.user, cantidad=0).count()
    valor_total_inventario = sum(
        (p.cantidad * (p.precio_costo or 0)) for p in productos
    )
    
    context = {
        'titulo': titulo,
        'productos': productos,
        'tipo_reporte': tipo_reporte,
        'total_productos': total_productos,
        'bajo_stock': bajo_stock,
        'sin_stock': sin_stock,
        'valor_total_inventario': valor_total_inventario,
    }
    return render(request, 'reportes/inventario.html', context)


# ==================== REPORTES FINANCIEROS ====================

@login_required
def reporte_financiero(request):
    """Reporte financiero: ingresos, gastos, utilidades"""
    from finanzas.models import CuentaPorCobrar, CuentaPorPagar
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Valores por defecto
    hoy = date.today()
    if not fecha_inicio:
        fecha_inicio = (hoy - timedelta(days=30)).isoformat()
    if not fecha_fin:
        fecha_fin = hoy.isoformat()
    
    # Ingresos por ventas (solo ventas activas, excluyendo anuladas)
    ingresos = Venta.objects.filter(
        owner=request.user,
        fecha_venta__gte=fecha_inicio,
        fecha_venta__lte=fecha_fin,
        estado='ACT'
    ).aggregate(total=Sum('total'))['total'] or 0
    
    # Gastos del período
    from gasto.models import Gasto
    gastos = Gasto.objects.filter(
        owner=request.user,
        fecha_gasto__gte=fecha_inicio,
        fecha_gasto__lte=fecha_fin,
        estado__in=['APROBADO', 'PAGADO']
    ).aggregate(total=Sum('monto'))['total'] or 0
    
    # Cuentas por cobrar pendientes
    from finanzas.models import CuentaPorCobrar, CuentaPorPagar
    cuentas_por_cobrar = CuentaPorCobrar.objects.filter(
        owner=request.user,
        estado__in=['PENDIENTE', 'PARCIAL']
    ).aggregate(total=Sum('saldo'))['total'] or 0
    
    # Cuentas por pagar pendientes
    cuentas_por_pagar = CuentaPorPagar.objects.filter(
        owner=request.user,
        estado__in=['PENDIENTE', 'PARCIAL']
    ).aggregate(total=Sum('saldo'))['total'] or 0
    
    # Utilidad neta
    utilidad_neta = ingresos - gastos
    
    context = {
        'titulo': 'Reporte Financiero',
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'ingresos': ingresos,
        'gastos': gastos,
        'utilidad_neta': utilidad_neta,
        'cuentas_por_cobrar': cuentas_por_cobrar,
        'cuentas_por_pagar': cuentas_por_pagar,
    }
    return render(request, 'reportes/financiero.html', context)


# ==================== REPORTES DE PRODUCTOS BAJO STOCK ====================

@login_required
def productos_bajo_stock(request):
    """Reporte de productos con bajo stock"""
    umbral_stock = 10
    productos = Producto.objects.filter(user=request.user, cantidad__lte=umbral_stock).order_by('cantidad')
    contexto = {
        'productos': productos,
        'umbral': umbral_stock
    }
    return render(request, 'reportes/productos_bajo_stock.html', contexto)


# ==================== EXPORTACIONES PDF/EXCEL ====================

@login_required
def exportar_ventas_excel(request):
    """Exportar reporte de ventas a Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecuta: pip install openpyxl', status=500)

    fecha_inicio = request.GET.get('fecha_inicio', (date.today() - timedelta(days=30)).isoformat())
    fecha_fin = request.GET.get('fecha_fin', date.today().isoformat())

    ventas = Venta.objects.filter(
        owner=request.user,
        fecha_venta__gte=fecha_inicio,
        fecha_venta__lte=fecha_fin
    ).order_by('-fecha_venta')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte de Ventas'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Reporte de Ventas ({fecha_inicio} a {fecha_fin})'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Encabezados
    headers = ['# Venta', 'Fecha', 'Cliente', 'Método Pago', 'Total', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Datos
    total_general = Decimal('0')
    for row, venta in enumerate(ventas, 4):
        ws.cell(row=row, column=1, value=venta.id_venta).border = border
        ws.cell(row=row, column=2, value=venta.fecha_venta.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row, column=3, value=str(venta.cliente) if venta.cliente else 'Consumidor Final').border = border
        ws.cell(row=row, column=4, value=getattr(venta, 'metodo_pago', 'efectivo')).border = border
        cell_total = ws.cell(row=row, column=5, value=float(venta.total))
        cell_total.number_format = '$#,##0.00'
        cell_total.border = border
        ws.cell(row=row, column=6, value=getattr(venta, 'estado', 'ACT')).border = border
        total_general += venta.total

    # Total
    fila_total = len(ventas) + 4
    ws.cell(row=fila_total, column=4, value='TOTAL:').font = Font(bold=True)
    cell_t = ws.cell(row=fila_total, column=5, value=float(total_general))
    cell_t.font = Font(bold=True)
    cell_t.number_format = '$#,##0.00'

    # Ajustar anchos
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ventas_{fecha_inicio}_{fecha_fin}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_inventario_excel(request):
    """Exportar reporte de inventario a Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse('openpyxl no está instalado.', status=500)

    productos = Producto.objects.filter(user=request.user).order_by('nombre')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    alert_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Reporte de Inventario - {date.today().strftime("%d/%m/%Y")}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Encabezados
    headers = ['Producto', 'Categoría', 'Stock', 'Precio Costo', 'Precio Venta', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Datos
    valor_total = Decimal('0')
    for row, p in enumerate(productos, 4):
        ws.cell(row=row, column=1, value=p.nombre).border = border
        ws.cell(row=row, column=2, value=str(p.categoria) if hasattr(p, 'categoria') and p.categoria else '-').border = border
        cell_stock = ws.cell(row=row, column=3, value=p.cantidad)
        cell_stock.border = border
        if p.cantidad <= 10:
            cell_stock.fill = alert_fill

        costo = float(p.precio_costo or 0)
        ws.cell(row=row, column=4, value=costo).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=float(p.precio)).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        val = p.cantidad * (p.precio_costo or 0)
        ws.cell(row=row, column=6, value=float(val)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        valor_total += val

    fila_total = len(productos) + 4
    ws.cell(row=fila_total, column=5, value='TOTAL:').font = Font(bold=True)
    cell_t = ws.cell(row=fila_total, column=6, value=float(valor_total))
    cell_t.font = Font(bold=True)
    cell_t.number_format = '$#,##0.00'

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="inventario_{date.today().isoformat()}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_ventas_pdf(request):
    """Exportar reporte de ventas a PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return HttpResponse('reportlab no está instalado. Ejecuta: pip install reportlab', status=500)

    fecha_inicio = request.GET.get('fecha_inicio', (date.today() - timedelta(days=30)).isoformat())
    fecha_fin = request.GET.get('fecha_fin', date.today().isoformat())

    ventas = Venta.objects.filter(
        owner=request.user,
        fecha_venta__gte=fecha_inicio,
        fecha_venta__lte=fecha_fin
    ).order_by('-fecha_venta')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ventas_{fecha_inicio}_{fecha_fin}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()

    elements = []

    # Título
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    elements.append(Paragraph(f'Reporte de Ventas', title_style))
    elements.append(Paragraph(f'Período: {fecha_inicio} a {fecha_fin}', styles['Normal']))
    elements.append(Spacer(1, 20))

    # Tabla
    data = [['# Venta', 'Fecha', 'Cliente', 'Método Pago', 'Total', 'Estado']]
    total_general = Decimal('0')
    for v in ventas:
        data.append([
            str(v.id_venta),
            v.fecha_venta.strftime('%d/%m/%Y %H:%M'),
            str(v.cliente) if v.cliente else 'Consumidor Final',
            getattr(v, 'metodo_pago', 'efectivo'),
            f'${v.total:,.2f}',
            getattr(v, 'estado', 'ACT'),
        ])
        total_general += v.total

    data.append(['', '', '', 'TOTAL:', f'${total_general:,.2f}', ''])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86C1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EBF5FB')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D5DBDB')),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f'Total de ventas: {len(ventas) - 0}', styles['Normal']))
    elements.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))

    doc.build(elements)
    return response


@login_required
def exportar_inventario_pdf(request):
    """Exportar reporte de inventario a PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return HttpResponse('reportlab no está instalado.', status=500)

    productos = Producto.objects.filter(user=request.user).order_by('nombre')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventario_{date.today().isoformat()}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()

    elements = []
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    elements.append(Paragraph('Reporte de Inventario', title_style))
    elements.append(Paragraph(f'Fecha: {date.today().strftime("%d/%m/%Y")}', styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [['Producto', 'Categoría', 'Stock', 'P. Costo', 'P. Venta', 'Valor Total']]
    valor_total = Decimal('0')
    for p in productos:
        val = p.cantidad * (p.precio_costo or 0)
        data.append([
            p.nombre[:30],
            str(p.categoria)[:20] if hasattr(p, 'categoria') and p.categoria else '-',
            str(p.cantidad),
            f'${p.precio_costo or 0:,.2f}',
            f'${p.precio:,.2f}',
            f'${val:,.2f}',
        ])
        valor_total += val

    data.append(['', '', '', '', 'TOTAL:', f'${valor_total:,.2f}'])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EAFAF1')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D5DBDB')),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f'Total de productos: {len(productos)}', styles['Normal']))
    elements.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))

    doc.build(elements)
    return response
